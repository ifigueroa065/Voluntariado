from tkinter import *
import os
from datetime import datetime
import webbrowser
from tkinter import messagebox
from tkinter import ttk
import tkinter.filedialog
import tkinter as tk
import openpyxl
from REPORTE import *
datos = [] #reporte
precios = [] #precios
preciosmq=[] #precios mq
subtotales = []


def CREAR_INTERFAZ():
    def DIALOGO():
        fd= tkinter.Tk()
        fd.withdraw()
        ruta=tkinter.filedialog.askopenfilename(
            initialdir="C:", 
            filetypes=(
                ("Libro de Excel", "*.xlsx"),
                ("Libro de Excel 97 a Excel 2003", "*.xls"),
                ("Todos los Archivos de Excel","*.*")
            ), 
            title = "ABRIR ARCHIVO"
        )
        if ruta=="":
            messagebox.showinfo(message="Debe cargar un archivo", title="ERROR")
        else:
            try:
                print("------> "+ ruta)
                rut.set("CARGA EXITOSA")
                book2 = openpyxl.load_workbook(ruta, data_only=True)
                celdas2 = book2.active
                for row in range(2,celdas2.max_row +1):
                    if(celdas2.cell(row,1).value is not None):
                        precios.append(Datos(celdas2.cell(row,1).value,celdas2.cell(row,2).value, celdas2.cell(row,3).value))
        
            finally:
                
                print("     **************************      ")
                print("            SUCCESSFULLY             ")
                print("     **************************      ")
    def DIALOGO2():
        fd= tkinter.Tk()
        fd.withdraw()
        ruta=tkinter.filedialog.askopenfilename(
            initialdir="C:", 
            filetypes=(
                ("Libro de Excel", "*.xlsx"),
                ("Libro de Excel 97 a Excel 2003", "*.xls"),
                ("Todos los Archivos de Excel","*.*")
            ), 
            title = "ABRIR ARCHIVO"
        )
        if ruta=="":
            messagebox.showinfo(message="Debe cargar un archivo", title="ERROR")
        else:
            try:
                print("------> "+ ruta)
                zm1.set("CARGA EXITOSA")
                book2 = openpyxl.load_workbook(ruta, data_only=True)
                celdas2 = book2.active
                for row in range(2,celdas2.max_row +1):
                    if(celdas2.cell(row,1).value is not None):
                        preciosmq.append(Datos(celdas2.cell(row,1).value,celdas2.cell(row,2).value, celdas2.cell(row,3).value))
        
            finally:
                
                print("     **************************      ")
                print("            SUCCESSFULLY             ")
                print("     **************************      ")

    def DIALOGO_REPORTE():
        TP=TIPO.get()
        fd= tkinter.Tk()
        fd.withdraw()
        ruta=tkinter.filedialog.askopenfilename(
            initialdir="C:", 
            filetypes=(
                ("Libro de Excel", "*.xlsx"),
                ("Libro de Excel 97 a Excel 2003", "*.xls"),
                ("Todos los Archivos de Excel","*.*")
            ), 
            title = "ABRIR ARCHIVO"
        )
        if ruta=="":
            messagebox.showinfo(message="Debe cargar un archivo", title="ERROR")
        else:
            try:
                print("------> "+ ruta)
                rut.set("CARGA EXITOSA")
                book = openpyxl.load_workbook(ruta, data_only=True)
                celdas = book.active
                for row in range(2,celdas.max_row):
                    if(celdas.cell(row,1).value is not None):
                        datos.append(Reporte(celdas.cell(row,1).value, celdas.cell(row,2).value, celdas.cell(row,3).value))
                
                if TP=="MQ":
                    print("--------------IMPRIMIENDO SUBTOTALES-------------")
                    x=0
                    contador=0
                    while x<len(datos):
                        for i in preciosmq:
                            if datos[x].nombre.upper().replace(" ", "")==i.nombre.upper().replace(" ", ""):
                                contador+=1
                                subtotal=datos[x].entregado_usuario*i.precio
                                print(str(contador)+ ")" +datos[x].nombre +"="+ str(subtotal))
                                subtotales.append(Subtotal(contador,datos[x].codigo,datos[x].nombre,datos[x].entregado_usuario,subtotal))
                                break
                        x+=1
                    print("----------------------------------------")
                    TOTAL=0
                    for i in subtotales:
                        TOTAL+=i.subtotal
                    print("TOTAL  = Q"+ str(TOTAL))
                else:
                    print("--------------IMPRIMIENDO SUBTOTALES-------------")
                    x=0
                    contador=0
                    while x<len(datos):
                        for i in precios:
                            if datos[x].nombre.upper().replace(" ", "")==i.nombre.upper().replace(" ", ""):
                                contador+=1
                                subtotal=datos[x].entregado_usuario*i.precio
                                print(str(contador)+ ")" +datos[x].nombre +"="+ str(subtotal))
                                subtotales.append(Subtotal(contador,datos[x].codigo,datos[x].nombre,datos[x].entregado_usuario,subtotal))
                                break
                        x+=1
                    print("----------------------------------------")
                    TOTAL=0
                    for i in subtotales:
                        TOTAL+=i.subtotal
                    print("TOTAL  = Q"+ str(TOTAL))
            finally:
                
                print("     **************************      ")
                print("            SUCCESSFULLY             ")
                print("     **************************      ")
    def VER_REPORTE():
        #obteniendo datos de inputs
        A=año.get()
        MO=Mes_inicial.get()
        M=Mes_final.get()

        DEPA=dpto.get()
        AR=area.get()
        MUN=municipio.get()
        TIPS=t_servicio.get()
        SERV=servicio.get()
        DIST=distrito.get()


        f = open('REPORTE.html','w', encoding="utf-8")
        f.write(""" 
                    <!DOCTYPE html>
                <html lang="en">

                <head>
            
            <meta charset="utf-8">
            <meta http-equiv="X-UA-Compatible" content="IE=edge">
            <meta name="viewport" content="width=device-width, initial-scale=1, shrink-to-fit=no">
            <meta name="description" content="">
            <meta name="author" content="">

            <title>ÁREA DE SALUD</title>
            <link href="img/icono.ico" rel="icon">
            <!-- Custom fonts for this template-->
            <link href="vendor/fontawesome-free/css/all.min.css" rel="stylesheet" type="text/css">
            <link
                href="https://fonts.googleapis.com/css?family=Nunito:200,200i,300,300i,400,400i,600,600i,700,700i,800,800i,900,900i"
                rel="stylesheet">

            <!-- Custom styles for this template-->
            
            <link href="css/sb-admin-2.min.css" rel="stylesheet">
            <link href="vendor/datatables/dataTables.bootstrap4.min.css" rel="stylesheet">

            </head>

            <body id="page-top">

            <!-- Page Wrapper -->
            <div id="wrapper">

            <!-- Sidebar -->
            <ul class="navbar-nav bg-gradient-primary sidebar sidebar-dark accordion" id="accordionSidebar">

                <!-- Sidebar - Brand -->
                <a class="sidebar-brand d-flex align-items-center justify-content-center" href="REPORTE.html">
                    <div class="sidebar-brand-icon rotate-n-15">
                        <i class="fas fa-laugh-wink"></i>
                    </div>
                    <div class="sidebar-brand-text mx-3">ANALISIS</div>
                </a>

            <!-- Divider -->
            <hr class="sidebar-divider my-0">

            <!-- Nav Item - Dashboard -->
            <li class="nav-item active">
                <a class="nav-link" href="REPORTE.html">
                    <i class="fas fa-bars"></i>
                    <span>REPORTE</span></a>
            </li>

            <!-- Divider -->
            <hr class="sidebar-divider">

            <!-- Heading -->
            <div class="sidebar-heading">
                OTROS
            </div>

            

            <!-- Nav Item - Utilities Collapse Menu -->
            <li class="nav-item">
                <a class="nav-link collapsed" href="#" data-toggle="collapse" data-target="#collapseUtilities"
                    aria-expanded="true" aria-controls="collapseUtilities">
                    <i class="fas fa-fw fa-2x"></i>
                    <span>BRESS</span>
                </a>
                
            </li>

           

        

            <!-- Divider -->
            <hr class="sidebar-divider d-none d-md-block">

            <!-- Sidebar Toggler (Sidebar) -->
            <div class="text-center d-none d-md-inline">
                <button class="rounded-circle border-0" id="sidebarToggle"></button>
            </div>

            </ul>
            <!-- End of Sidebar -->

            <!-- Content Wrapper -->
            <div id="content-wrapper" class="d-flex flex-column">

                <!-- Main Content -->
                <div id="content">

                    <!-- Topbar -->
                    <nav class="navbar navbar-expand navbar-light bg-white topbar mb-4 static-top shadow">

                        <!-- Sidebar Toggle (Topbar) -->
                        <button id="sidebarToggleTop" class="btn btn-link d-md-none rounded-circle mr-3">
                            <i class="fa fa-bars"></i>
                        </button>


                        <!-- Topbar Navbar -->
                        <ul class="navbar-nav ml-auto">

                            <!-- Nav Item - Search Dropdown (Visible Only XS) -->
                            <li class="nav-item dropdown no-arrow d-sm-none">
                                <a class="nav-link dropdown-toggle" href="#" id="searchDropdown" role="button"
                                    data-toggle="dropdown" aria-haspopup="true" aria-expanded="false">
                                    <i class="fas fa-search fa-fw"></i>
                                </a>
                                
                            </li>

                            

                            

                            <div class="topbar-divider d-none d-sm-block"></div>

                            <!-- Nav Item - User Information -->
                            <li class="nav-item dropdown no-arrow">
                                <a class="nav-link dropdown-toggle" href="#" id="userDropdown" role="button"
                                    data-toggle="dropdown" aria-haspopup="true" aria-expanded="false">
                                    <span class="mr-2 d-none d-lg-inline text-gray-600 small">Administrador</span>
                                    <img class="img-profile rounded-circle"
                                        src="img/undraw_profile.svg">
                                </a>
                            </li>

                        </ul>

                    </nav>
                    <!-- End of Topbar -->

                    <!-- Begin Page Content -->
                    <div class="container-fluid">

                        <!-- Page Heading -->
                        <div class="d-sm-flex align-items-center justify-content-between mb-4">
                            <h1 class="h3 mb-0 text-gray-800">ÁREA DE SALUD DE CHIMALTENANGO</h1>
                            <a href="#" class="d-none d-sm-inline-block btn btn-sm btn-primary shadow-sm"><i
                                    class="fas fa-download fa-sm text-white-50"></i> Descargar Reporte</a>
                        </div>

                        <!-- Content Row -->
                        <div class="row">       

                                                <!-- Earnings (Monthly) Card Example -->
                        <div class="col-xl-3 col-md-6 mb-4">
                            <div class="card border-left-primary shadow h-100 py-2">
                                <div class="card-body">
                                    <div class="row no-gutters align-items-center">
                                        <div class="col mr-2">
                                            <div class="text-xs font-weight-bold text-primary text-uppercase mb-1">
                                                Departamento</div>
                                            <div class="h5 mb-0 font-weight-bold text-gray-800"> 
        """)
        f.write(DEPA) #DEPARTAMENTO
        f.write("""
            </div>
                                            </div>
                                            <div class="col-auto">
                                                <i class="fas fa-fw"></i>
                                            </div>
                                        </div>
                                    </div>
                                </div>
                            </div>

                            <!-- Earnings (Monthly) Card Example -->
                            <div class="col-xl-3 col-md-6 mb-4">
                                <div class="card border-left-success shadow h-100 py-2">
                                    <div class="card-body">
                                        <div class="row no-gutters align-items-center">
                                            <div class="col mr-2">
                                                <div class="text-xs font-weight-bold text-success text-uppercase mb-1">
                                                    Distrito</div>
                                                <div class="h5 mb-0 font-weight-bold text-gray-800">

        """)
        f.write(DIST) #DISTRITO
        f.write("""
                </div>
                                        </div>
                                        <div class="col-auto">
                                            <i class="fas fa-fw"></i>
                                        </div>
                                    </div>
                                </div>
                            </div>
                        </div>

                        <!-- Earnings (Monthly) Card Example -->
                        <div class="col-xl-3 col-md-6 mb-4">
                            <div class="card border-left-info shadow h-100 py-2">
                                <div class="card-body">
                                    <div class="row no-gutters align-items-center">
                                        <div class="col mr-2">
                                            <div class="text-xs font-weight-bold text-info text-uppercase mb-1">Del Mes
                                            </div>
                                            <div class="row no-gutters align-items-center">
                                                <div class="col-auto">
                                                    <div class="h5 mb-0 mr-3 font-weight-bold text-gray-800">
        """)
        f.write(MO) #MES INICIAL
        f.write("""
                        </div>
                                                </div>
                                                
                                            </div>
                                        </div>
                                        <div class="col-auto">
                                            <i class="fas fa-calendar fa-2x text-gray-300"></i>
                                        </div>
                                    </div>
                                </div>
                            </div>
                        </div>

                        <!-- Pending Requests Card Example -->
                        <div class="col-xl-3 col-md-6 mb-4">
                            <div class="card border-left-warning shadow h-100 py-2">
                                <div class="card-body">
                                    <div class="row no-gutters align-items-center">
                                        <div class="col mr-2">
                                            <div class="text-xs font-weight-bold text-warning text-uppercase mb-1">
                                                Al mes</div>
                                            <div class="h5 mb-0 font-weight-bold text-gray-800">
        """)
        f.write(M) #MES FINAL
        f.write("""
                    </div>
                                        </div>
                                        <div class="col-auto">
                                            <i class="fas fa-calendar fa-2x text-gray-300"></i>
                                        </div>
                                    </div>
                                </div>
                            </div>
                        </div>
                    </div>

                    <!-- Content Row -->

                    <div class="row">

                        

                        

                    
                    <!-- TABLA RESUMEN-->
                    <h1 class="h3 mb-2 text-gray-800">
        """)
        f.write(MUN) # MUNICIPIO
        f.write("""
                </h1>
                    <p class="mb-4">Reporte de Balance, Requisición y Envío de Suministros</p>

                    <!-- TABLA DE MEDICAMENTOS Y MÉDIDO QUIRURGICO -->
                    <div class="card shadow mb-4">
                        <div class="card-header py-3">
                            <h6 class="m-0 font-weight-bold text-primary">
        
        """)
        f.write(TIPS) #TIPO DE SERVICIO
        f.write("""
                        </h6>
                        </div>
                        <div class="card-body">
                            <div class="table-responsive">
                                <table class="table table-bordered" id="dataTable" width="100%" cellspacing="0">
                                    <thead>
                                        <tr>
                                            <th>Número de orden</th>
                                            <th>Código</th>
                                            <th>Descripción de Articulo/Producto</th>
                                            <th>Unidad de Medida</th>
                                            <th>Cantidad Autorizada</th>
                                            <th>Cantidad despachada</th>
                                            <th>Subtotal</th>
                                        </tr>
                                    </thead>
                                    <tfoot>
                                        <th>Número de orden</th>
                                            <th>Código</th>
                                            <th>Descripción de Articulo/Producto</th>
                                            <th>Unidad de Medida</th>
                                            <th>Cantidad Autorizada</th>
                                            <th>Cantidad despachada</th>
                                            <th>Subtotal </th>
                                    </tfoot>
                                    <tbody>
        """)

        for i in subtotales:
            p="{0:.2f}".format(float(i.subtotal))
            f.write("<tr>")
            f.write(" <td><center>"+str(i.id)+"</center></td>"
                    +"<td><center>"+str(i.codigo)+"</center></td>"
                    +"<td><center>"+str(i.nombre)+"</center></td>"
                    +"<td><center>"+"x"+"</center></td>"
                    +"<td><center>"+str(i.entregado)+"</center></td>"
                    +"<td><center>"+str(i.entregado)+"</center></td>"
                    +"<td><center>"+ "Q"+str(p)+"</center></td>"
            )     
            f.write("<t/r>")

        f.write("""
                                </tbody>
                                            </table>
                                        </div>
                                    </div>
                                </div>
                                <!-- Content Row -->
                                <div class="row">

                                    <!-- Content Column -->

                                    <div class="col-auto">

                                        


                                    </div>
                                </div>
                            </div>
                            <!-- /.container-fluid -->

                        </div>
                        <!-- End of Main Content -->

                        <!-- Footer -->
                        <footer class="sticky-footer bg-white">
                            <div class="container my-auto">
                                <div class="copyright text-center my-auto">
                                    <span>&copy; Facultad de Ingeniería 2021</span>
                                </div>
                            </div>
                        </footer>
                        <!-- End of Footer -->

                    </div>
                    <!-- End of Content Wrapper -->

                </div>
                <!-- End of Page Wrapper -->

                <!-- Scroll to Top Button-->
                <a class="scroll-to-top rounded" href="#page-top">
                    <i class="fas fa-angle-up"></i>
                </a>

                <!-- Logout Modal-->
                <div class="modal fade" id="logoutModal" tabindex="-1" role="dialog" aria-labelledby="exampleModalLabel"
                    aria-hidden="true">
                    <div class="modal-dialog" role="document">
                        <div class="modal-content">
                            <div class="modal-header">
                                <h5 class="modal-title" id="exampleModalLabel">Ready to Leave?</h5>
                                <button class="close" type="button" data-dismiss="modal" aria-label="Close">
                                    <span aria-hidden="true">×</span>
                                </button>
                            </div>
                            <div class="modal-body">Select "Logout" below if you are ready to end your current session.</div>
                            <div class="modal-footer">
                                <button class="btn btn-secondary" type="button" data-dismiss="modal">Cancel</button>
                                <a class="btn btn-primary" href="login.html">Logout</a>
                            </div>
                        </div>
                    </div>
                </div>

                <!-- Bootstrap core JavaScript-->
                <script src="vendor/jquery/jquery.min.js"></script>
                <script src="vendor/bootstrap/js/bootstrap.bundle.min.js"></script>

                <!-- Core plugin JavaScript-->
                <script src="vendor/jquery-easing/jquery.easing.min.js"></script>

                <!-- Custom scripts for all pages-->
                <script src="js/sb-admin-2.min.js"></script>

                <!-- Page level plugins -->
                <script src="vendor/chart.js/Chart.min.js"></script>

                <!-- Page level custom scripts -->
                <script src="js/demo/chart-area-demo.js"></script>
                <script src="js/demo/chart-pie-demo.js"></script>

                <!-- Page level plugins -->
                <script src="vendor/datatables/jquery.dataTables.min.js"></script>
                <script src="vendor/datatables/dataTables.bootstrap4.min.js"></script>
            
                <!-- Page level custom scripts -->
                <script src="js/demo/datatables-demo.js"></script>

            </body>

            </html>
                    
        
        """)


        f.close()
        
        webbrowser.open_new_tab('REPORTE.html')
        


    #--------------CREANDO VENTANA PRINCIPAL--------------
    root=Tk()
    root.title("VOLUNTARIADO")
    root.iconbitmap('img\icono.ico')

    rut=StringVar()
    zm1=StringVar()

    nt=ttk.Notebook(root)
    nt.pack(fill="both",expand="yes")

    s = ttk.Style()
    # Create style used by default for all Frames
    s.configure('TFrame', background='#1F618D')

    #--------------FRAME INICIO--------------
    s.configure('Frame1.TFrame', background='#1F618D')
    V1 = ttk.Frame(nt, style='Frame1.TFrame')
    nt.add(V1, text="INICIO")

    #--------------FRAME CARGAR ARCHIVOS--------------
    s.configure('Frame2.TFrame', background='#1F618D')
    V2 = ttk.Frame(nt, style='Frame2.TFrame')
    nt.add(V2, text="PRECIOS")

    Label(V2,textvariable=rut,font="Helvetica 16",bg="#1F618D").place(x=100,y=280)
    rut.set("NO SE HA CARGADO NADA")
    Button(V2,text="SELECCIONAR ARCHIVO",command=DIALOGO,font="Helvetica 12",height=5,width=25).place(x=120, y=110)
   
    Label(V2,textvariable=zm1,font="Helvetica 16",bg="#1F618D").place(x=560,y=280)
    zm1.set("NO SE HA CARGADO NADA")
    Button(V2,text="SELECCIONAR ARCHIVO",command=DIALOGO2,font="Helvetica 12",height=5,width=25).place(x=520, y=110)

    L1=StringVar()

    l2=StringVar()
    l3=StringVar()

    xo=IntVar()
    yo=IntVar()
    
  
    Label(V2,textvariable=L1,font="Helvetica 16",bg="#1F618D").place(x=30,y=30)
    L1.set("CARGAR ARCHIVO DE PRECIOS (MED)")


    Label(V2,textvariable=l2,font="Helvetica 16",bg="#1F618D").place(x=500,y=30)
    l2.set("CARGAR ARCHIVO DE PRECIOS (MQ)")
   

    #--------------FRAME REPORTES--------------
    s.configure('Frame3.TFrame', background='#1F618D')
    V3 = ttk.Frame(nt, style='Frame3.TFrame')
    nt.add(V3, text=" VISUALIZAR REPORTE")

    icodoct=PhotoImage(file="img\doct.png")
    icodoct.subsample(1,1)
    #Button(V3,image=icodoct,font="Helvetica 14",width=300,height=300).place(x=100, y=130)
    
    Label(V3,textvariable=rut,font="Helvetica 16",bg="#1F618D").place(x=150,y=400)
    rut.set("NO SE HA CARGADO NADA")
    Button(V3,text="SELECCIONAR ARCHIVO",command=DIALOGO_REPORTE,font="Helvetica 12").place(x=250, y=350)
    Button(V3,text="VER REPORTE",command=VER_REPORTE,height=5,width=25,font="Helvetica 12").place(x=650, y=350)
    
    L6=StringVar()

    año=StringVar()



    dpto=StringVar()
    area=StringVar()
    distrito=StringVar()
    municipio=StringVar()
    t_servicio=StringVar()
    servicio=StringVar()


    l9=StringVar()
    l8=StringVar()
    l7=StringVar()
    l6=StringVar()
    l5=StringVar()
    l4=StringVar()
    a=StringVar()
    b=StringVar()
    c=StringVar()
    
  
    Label(V3,textvariable=L6,font="Helvetica 16",bg="#1F618D").place(x=70,y=30)
    L6.set("DATOS PARA EL REPORTE")

    Label(V3,textvariable=l9,font="Helvetica 12",bg="#1F618D",fg="white").place(x=75,y=140)
    l9.set("Departamento")
    """Label(V3,textvariable=l8,font="Helvetica 12",bg="#1F618D",fg="white").place(x=75,y=180)
    l8.set("Area")
    Label(V3,textvariable=l7,font="Helvetica 12",bg="#1F618D",fg="white").place(x=75,y=220)
    l7.set("Distrito")"""
    Label(V3,textvariable=l6,font="Helvetica 12",bg="#1F618D",fg="white").place(x=75,y=180)
    l6.set("Municipio")
    Label(V3,textvariable=l5,font="Helvetica 12",bg="#1F618D",fg="white").place(x=475,y=180)
    """l5.set("Tipo de Servicio")
    Label(V3,textvariable=l4,font="Helvetica 12",bg="#1F618D",fg="white").place(x=475,y=220)
    l4.set("Servicio")"""
    
    Label(V3,textvariable=a,font="Helvetica 12",bg="#1F618D",fg="white").place(x=450,y=40)
    a.set("Año")
    Label(V3,textvariable=b,font="Helvetica 12",bg="#1F618D",fg="white").place(x=570,y=40)
    b.set("Del Mes")
    Label(V3,textvariable=c,font="Helvetica 12",bg="#1F618D",fg="white").place(x=760,y=40)
    c.set("Al mes")

    Entry(V3,textvariable=año,font="Helvetica 11",width=5).place(x=500,y=40)
    #Entry(V3,textvariable=Mes_inicial,font="Helvetica 11",width=10).place(x=650,y=40)
    #Entry(V3,textvariable=Mes_final,font="Helvetica 11",width=10).place(x=820,y=40)
    Mes_inicial=ttk.Combobox(V3,width=10,font="Helvetica 11",state="readonly")
    Mes_inicial.place(x=650,y=40)
    Mes_inicial['values']=('Enero','Febrero','Marzo ','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre')

    Mes_final=ttk.Combobox(V3,width=10,font="Helvetica 11",state="readonly")
    Mes_final.place(x=820,y=40)
    Mes_final['values']=('Enero','Febrero','Marzo ','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre')


    TIPO=ttk.Combobox(V3,width=10,font="Helvetica 14",state="readonly")
    TIPO.place(x=100,y=350)
    TIPO['values']=('MED','MQ')
    
    Entry(V3,textvariable=dpto,font="Helvetica 12").place(x=200,y=140)
    #Entry(V3,textvariable=distrito,font="Helvetica 12").place(x=200,y=180)
    #Entry(V3,textvariable=t_servicio,font="Helvetica 12").place(x=200,y=220)

    Entry(V3,textvariable=distrito,font="Helvetica 12").place(x=200,y=180)
    #Entry(V3,textvariable=municipio,font="Helvetica 12").place(x=600,y=180)
    #Entry(V3,textvariable=servicio,font="Helvetica 12").place(x=600,y=220)
    

    root.geometry("950x550")
    root.mainloop()
    
    
    
CREAR_INTERFAZ()