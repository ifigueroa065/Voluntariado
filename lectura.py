import openpyxl

#PROBANDO LECTURA DE XLS
try:
    book=openpyxl.load_workbook('DATA\PMDMQ.xlsx',data_only=True)

    #Fijando Hoja
    hoja=book.active
    
    print("------------------------LECTURA------------------------")
    for fila in hoja.iter_rows(min_col=hoja.min_column,max_col=hoja.max_column):
        mex=[celda.value for celda in fila]
        print(mex)
finally:
    print("________________________________________________________")
    print("                                                        ")
    print("                     SUCESSFULLY                        ")
    print("________________________________________________________")