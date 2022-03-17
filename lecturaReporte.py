import openpyxl
import os

book = openpyxl.load_workbook('DATA/reporte2Data.xlsx', data_only=True)
book2 = openpyxl.load_workbook('DATA/PRECIOS.xlsx', data_only=True)


celdas = book.active
celdas2 = book2.active

class Subtotal():
    def __init__(self, id, codigo, nombre, entregado, subtotal):
        self.id=id
        self.nombre=nombre
        self.entregado=entregado
        self.codigo = codigo
        self.subtotal = subtotal


        
class Datos():
    def __init__(self, codigo, nombre, precio):
        self.codigo = codigo
        self.nombre = nombre
        self.precio = precio

        

class Reporte():
    def __init__(self, codigo, nombre, entregado_usuario):
        self.codigo = codigo
        self.nombre = nombre
        self.entregado_usuario = entregado_usuario

datos = [] #reporte
precios = [] #precios
subtotales = []

os.system('cls')

#LEYENDO ARCHIVO DEL MODULO
for row in range(2,celdas.max_row):
    if(celdas.cell(row,1).value is not None):
        datos.append(Reporte(celdas.cell(row,1).value, celdas.cell(row,2).value, celdas.cell(row,3).value))

"""print("----------------------DATOS DEL REPORTE-------------")
for i in datos:
    print(i.nombre + "=" + str(i.entregado_usuario))"""

#LEYENDO ARCHIVO DE PRECIOS
for row in range(2,celdas2.max_row +1):
    if(celdas2.cell(row,1).value is not None):
        precios.append(Datos(celdas2.cell(row,1).value,celdas2.cell(row,2).value, celdas2.cell(row,3).value))

"""print("----------------------PRECIOS-------------")
for i in precios:
    print(i.nombre + "=" + str(i.precio))"""

print("--------------IMPRIMIENDO SUBTOTALES-------------")
x=0
contador=0
while x<len(datos):
    for i in precios:
        if datos[x].nombre.upper().replace(" ", "")==i.nombre.upper().replace(" ", ""):
            contador+=1
            subtotal=datos[x].entregado_usuario*i.precio
            print(str(contador)+ ")" +datos[x].nombre +"="+ str(subtotal))
            subtotales.append(Subtotal(contador,datos[x].codigo,datos[x].nombre,datos[x].entregado_usuario,subtotal))
            break
    x+=1
print("----------------------------------------")
TOTAL=0
for i in subtotales:
    TOTAL+=i.subtotal
print("TOTAL  = Q"+ str(TOTAL))